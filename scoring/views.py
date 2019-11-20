from django.shortcuts import render
from django.views.generic import ListView
from .models import *
import csv
from openpyxl import load_workbook
from string import ascii_uppercase
from .convertStrToNum import float1, int1
from django.http import HttpResponse
from .views import *
from statistics import stdev, mean

# Create your views here.
class HomeListView(ListView):
    model = Judge
    template_name = 'home.html'

def home(request):
    return render(request, 'home.html')
    #return render(ListView, 'home.html')

    #def import_db(request):
    #    f = open('Open Scoring UNH Ver09.csv', 'r')
    #    for line in f:
    #        line = line.split(',')
    #        tmp = Judge.object.create()
    #        tmp.judge_id = line[]

def display_judges(request):
    items = Judge.objects.all()
    button = "judges"
    context = {
        'button' : button,
        'items' : items,
        #'name' :
    }
    return render(request, 'home.html', context)

def display_projects(request):
    button = "projects"
    project_items = Project.objects.all()
    context = {
        'button' : button,
        'project_items' : project_items,
        #'name' :
    }
    return render(request, 'home.html', context)

def display_judge_assignments(request):
    button = "judge_assignments"
    ja_items = Judge_Assignment.objects.all()
    context = {
        'button' : button,
        'ja_items' : ja_items,
        #'name' :
    }
    return render(request, 'home.html', context)

def display_students(request):
    button = "students"
    student_items = Student.objects.all()
    context = {
        'button' : button,
        'student_items' : student_items,
        #'name' :
    }
    return render(request, 'home.html', context)

def import_file(request):
    if request.method=='POST':
        file = request.FILES['document']
        import_data(file)
    return render(request, 'import.html')

def import_data(request):
    wb = load_workbook(filename = request, data_only=True)
    sheet = wb.active
    import_judge(sheet)
    import_project(sheet)
    import_student(sheet)
    import_judge_assignment(sheet)

def import_judge(sheet):
    alphabet = ascii_uppercase
    #JUDGE
    judge_id = []
    name = []
    for C in alphabet[13:26]:
        j_id = sheet[str(C)+str(4)].value
        j_n = sheet[str(C)+str(3)].value
        if '-' in j_id and j_n:
            continue
        if j_id not in judge_id:
            judge_id.append(j_id)
        if j_n not in name:
            name.append(j_n)
    for C1 in alphabet[0:8]:
        for C2 in alphabet:
            j_id = sheet[str(C1)+str(C2)+str(4)].value
            j_n = sheet[str(C1)+str(C2)+str(3)].value
            if j_id is None:
                break
            if '-' in j_id and j_n:
                continue
            if j_id not in judge_id:
                judge_id.append(j_id)
            if j_n not in name:
                name.append(j_n)

    #SAVE JUDGE
    for j_id, j_n in zip(judge_id, name):
        j = Judge(j_id, j_n)
        j.save()

def import_project(sheet):
    for s in range(5, 71):
    #PROJECT
        project_id = sheet['D'+str(s)].value
        description = sheet['L'+str(s)].value
        project_title = sheet['J'+str(s)].value
        project_category = sheet['K'+str(s)].value
        avg_score = None #sheet['LB'+str(s)].value
        rank = None #sheet['LC'+str(s)].value
        z_score = None #sheet['LD'+str(s)].value
        z_score_rank = None
        avg_01 = None
        avg_01_rank = None
        scaled_score = sheet['LH'+str(s)].value
        scaled_rank = sheet['LJ'+str(s)].value
        scaled_z = sheet['LI'+str(s)].value
        isef_score = sheet['LK'+str(s)].value
        isef_rank = sheet['LL'+str(s)].value

    #SAVE PROJECT
        p = Project(project_id, description, project_title, project_category, float1(avg_score), int1(rank), float1(z_score), z_score_rank, avg_01, avg_01_rank, float1(scaled_score), float1(scaled_rank), float1(scaled_z), float1(isef_score), int1(isef_rank))
        p.save()

def import_student(sheet):
    for s in range(5, 71):
    #STUDENT
        project_id = sheet['D'+str(s)].value
        ids = [sheet['E'+str(s)].value, sheet['F'+str(s)].value, sheet['G'+str(s)].value]
        school = sheet['H'+str(s)].value

    #SAVE STUDENT
        for id in ids:
            if id is not None:
                s = Student(id, school, project_id)
                s.save()

def import_judge_assignment(sheet):
    #JUDGE ASSIGNMENT
    goal_score = None
    plan_score = None
    action_score = None
    result_analysis_score = None
    communication_score = None
    # raw_score = goal_score + plan_score + action_score + result_analysis_score + communication_score

    #SAVE JUDGE_ASSIGNMENT
    id = 0
    for row in sheet.iter_rows(min_row=5, min_col=13, max_row=71, max_col=220):
        for cell in row:
            if isinstance(cell.value,int):
                id = id + 1
                j_id = sheet.cell(row = 4, column = cell.column).value
                if '-' in j_id:
                    continue
                p_id = sheet.cell(row = cell.row, column = 4).value
                raw_score = cell.value
                ja = Judge_Assignment(id, j_id, p_id, goal_score, plan_score, action_score, result_analysis_score, communication_score, raw_score)
                ja.save()

def export_jugde_assignment(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="judge_assignments.csv"'

    writer = csv.writer(response)
    writer.writerow(['Judge ID', 'Project ID', 'Raw Score'])

    jas = Judge_Assignment.objects.all().values_list('judge_id', 'project_id', 'raw_score')
    for ja in jas:
        writer.writerow(list(ja))
    return response

def remove_all_data(request):
    Judge.objects.all().delete()
    Project.objects.all().delete()
    Student.objects.all().delete()
    Judge_Assignment.objects.all().delete()
    return render(request, 'home.html')

def cal_average_score():
    done_list = []
    jas = list(Judge_Assignment.objects.all())
    for ja in jas:
        if ja.project_id not in done_list:
            score = 0
            done_list.append(ja.project_id)
            project = Judge_Assignment.objects.filter(project_id = ja.project_id.project_id)
            for p in project:
                score = score + p.raw_score
            avg_score = score / len(list(project))
            if score == 0:
                continue
            project = Project.objects.get(project_id = ja.project_id.project_id)
            project.avg_score = avg_score
            project.save()

def sort_rank():
    projects = list(Project.objects.all().order_by('-avg_score'))
    for i in range(len(projects)):
        projects[i].rank = i+1
        if not isinstance(projects[i].avg_score,float):
            projects[i].rank = None
            continue
        if i == 0:
            projects[i].rank = i+1
            continue
        if projects[i].avg_score == projects[i-1].avg_score:
            projects[i].rank = projects[i-1].rank
            continue
        projects[i].save()

def cal_z_score():
    jas = list(Judge_Assignment.objects.all())
    for ja in jas:
        jas_projects = list(Judge_Assignment.objects.filter(judge_id = ja.judge_id))
        all_judge_scores = []
        for jas_project in jas_projects:
            all_judge_scores.append(jas_project.raw_score)
        if not all_judge_scores:
            continue
        ja.z_score = (ja.raw_score - mean(all_judge_scores))/stdev(all_judge_scores)
        ja.save()

def sort_judge_rank():
    judges = list(Judge.objects.all())
    for judge in judges:
        jas = list(Judge_Assignment.objects.filter(judge_id = judge.judge_id).order_by('-raw_score'))
        for index in range(len(jas)):
            if jas[index].raw_score is None:
                continue
            jas[index].rank = (1-(1/(len(jas)-1))*index)
            if index > 0 and (jas[index].raw_score == jas[index-1].raw_score):
                jas[index].rank = jas[index-1].rank
            jas[index].save()

def cal_avg_z_score():
    projects = list(Project.objects.all())
    for project in projects:
        jas_z_score = list(Judge_Assignment.objects.filter(project_id = project.project_id).values_list('z_score', flat=True))
        if not jas_z_score:
            continue
        project.z_score = mean(jas_z_score)
        project.save()

def sort_z_score_rank():
    projects = list(Project.objects.all().order_by('-z_score'))
    for index in range(len(projects)):
        if projects[index].z_score is None:
            continue
        projects[index].z_score_rank = index+1
        if projects[index].z_score == projects[index-1].z_score:
            projects[index].z_score_rank = projects[index-1].z_score_rank
        projects[index].save()

def cal_avg_01():
    projects = list(Project.objects.all())
    for project in projects:
        jas = Judge_Assignment.objects.filter(project_id = project.project_id)
        all_judge_z_ranks = []
        for ja in jas:
            all_judge_z_ranks.append(ja.rank)
        if not all_judge_z_ranks:
            continue
        project.avg_01 = mean(all_judge_z_ranks)
        project.save()

def sort_avg_01_rank():
    projects = list(Project.objects.all().order_by('-avg_01'))
    for index in range(len(projects)):
        if projects[index].avg_01 is None:
            continue
        projects[index].avg_01_rank = index+1
        if projects[index].avg_01 == projects[index-1].avg_01:
            projects[index].avg_01_rank = projects[index-1].avg_01_rank
        projects[index].save()

def cal_scaled_score():
    projects = list(Project.objects.all())
    for index in range(len(projects)):
        min = list(Project.objects.all().order_by('avg_score'))[index].avg_score
        if min is not None:
            break
    max = list(Project.objects.all().order_by('-avg_score'))[0].avg_score
    rangevalue = max - min
    for project in projects:
        if project.avg_score is None:
            continue
        project.scaled_score = ((project.avg_score - min)/rangevalue)*25 + 25
        project.save()

def calculate_scores(request):
    cal_average_score()
    sort_rank()
    cal_z_score()
    sort_judge_rank()
    cal_avg_z_score()
    sort_z_score_rank()
    cal_avg_01()
    sort_avg_01_rank()
    cal_scaled_score()
    return render(request, 'home.html')
